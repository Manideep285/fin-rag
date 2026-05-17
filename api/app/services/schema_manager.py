"""Schema Manager — Auto-introspect & cache schemas from structured data sources.

Handles:
- CSV files → infer column names/types, load into DuckDB
- Excel files → sheet-level schema introspection, load into DuckDB
- External SQL databases → introspect tables and columns via DSN
- Schema caching for fast LLM prompting

Each project gets its own DuckDB namespace (schema) for isolation.
"""
from __future__ import annotations
import csv
import hashlib
import json
from functools import lru_cache
from io import BytesIO, StringIO
from pathlib import Path
from typing import Optional
from uuid import UUID

from ..config import settings
from ..logging_setup import log

# Lazy import — DuckDB is optional for the api container
_duckdb_conn = None


def _get_duckdb():
    """Lazy singleton DuckDB connection."""
    global _duckdb_conn
    if _duckdb_conn is None:
        import duckdb

        db_path = settings.duckdb_path
        Path(db_path).parent.mkdir(parents=True, exist_ok=True)
        _duckdb_conn = duckdb.connect(db_path)
        log.info("schema_manager.duckdb_connected", path=db_path)
    return _duckdb_conn


def _sanitize_name(name: str) -> str:
    """Sanitize a name for use as SQL identifier."""
    import re
    cleaned = re.sub(r"[^a-zA-Z0-9_]", "_", name)
    if cleaned[0:1].isdigit():
        cleaned = "t_" + cleaned
    return cleaned.lower()[:63]


def _project_schema(project_id: UUID) -> str:
    """Get the DuckDB schema name for a project."""
    return f"p_{str(project_id).replace('-', '_')}"


# ---------------------------------------------------------------------------
# CSV → DuckDB ingestion
# ---------------------------------------------------------------------------

def ingest_csv_to_duckdb(
    project_id: UUID,
    source_id: UUID,
    source_name: str,
    raw: bytes,
) -> dict:
    """Load a CSV file into DuckDB and return schema information.

    Returns:
        {
            "table_name": str,
            "schema_name": str,
            "columns": [{"name": str, "type": str}],
            "row_count": int,
        }
    """
    conn = _get_duckdb()
    schema = _project_schema(project_id)
    table = _sanitize_name(Path(source_name).stem)
    full_table = f"{schema}.{table}"

    # Create schema if not exists
    conn.execute(f"CREATE SCHEMA IF NOT EXISTS {schema}")

    # Read CSV into DuckDB directly
    text = raw.decode("utf-8", errors="replace")
    reader = csv.reader(StringIO(text))
    rows = [r for r in reader if any(c.strip() for c in r)]

    if not rows:
        return {"table_name": table, "schema_name": schema, "columns": [], "row_count": 0}

    headers = [_sanitize_name(h) or f"col_{i}" for i, h in enumerate(rows[0])]
    data_rows = rows[1:]

    # Infer types from first 100 rows
    col_types = _infer_column_types(headers, data_rows[:100])

    # Create table
    col_defs = ", ".join(f'"{h}" {t}' for h, t in zip(headers, col_types))
    conn.execute(f"DROP TABLE IF EXISTS {full_table}")
    conn.execute(f"CREATE TABLE {full_table} ({col_defs})")

    # Insert data
    if data_rows:
        placeholders = ", ".join(["?"] * len(headers))
        for row in data_rows:
            # Pad or trim to match header count
            padded = (row + [""] * len(headers))[:len(headers)]
            values = [_cast_value(v, t) for v, t in zip(padded, col_types)]
            conn.execute(f"INSERT INTO {full_table} VALUES ({placeholders})", values)

    row_count = conn.execute(f"SELECT COUNT(*) FROM {full_table}").fetchone()[0]

    columns = [{"name": h, "type": t} for h, t in zip(headers, col_types)]
    log.info("schema_manager.csv_ingested", table=full_table, rows=row_count, cols=len(columns))

    return {
        "table_name": table,
        "schema_name": schema,
        "columns": columns,
        "row_count": row_count,
    }


# ---------------------------------------------------------------------------
# Excel → DuckDB ingestion
# ---------------------------------------------------------------------------

def ingest_xlsx_to_duckdb(
    project_id: UUID,
    source_id: UUID,
    source_name: str,
    raw: bytes,
) -> list[dict]:
    """Load an Excel file into DuckDB (one table per sheet).

    Returns list of schema info dicts, one per sheet.
    """
    from openpyxl import load_workbook

    conn = _get_duckdb()
    schema = _project_schema(project_id)
    conn.execute(f"CREATE SCHEMA IF NOT EXISTS {schema}")

    wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    results = []

    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        rows = [list(r) for r in rows_iter]
        rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]

        if not rows:
            continue

        sheet_table = _sanitize_name(f"{Path(source_name).stem}_{ws.title}")
        full_table = f"{schema}.{sheet_table}"

        headers = [
            _sanitize_name(str(h) if h else f"col_{i}") or f"col_{i}"
            for i, h in enumerate(rows[0])
        ]
        data_rows = [[str(c) if c is not None else "" for c in r] for r in rows[1:]]

        col_types = _infer_column_types(headers, data_rows[:100])

        col_defs = ", ".join(f'"{h}" {t}' for h, t in zip(headers, col_types))
        conn.execute(f"DROP TABLE IF EXISTS {full_table}")
        conn.execute(f"CREATE TABLE {full_table} ({col_defs})")

        if data_rows:
            placeholders = ", ".join(["?"] * len(headers))
            for row in data_rows:
                padded = (row + [""] * len(headers))[:len(headers)]
                values = [_cast_value(v, t) for v, t in zip(padded, col_types)]
                conn.execute(f"INSERT INTO {full_table} VALUES ({placeholders})", values)

        row_count = conn.execute(f"SELECT COUNT(*) FROM {full_table}").fetchone()[0]

        columns = [{"name": h, "type": t} for h, t in zip(headers, col_types)]
        results.append({
            "table_name": sheet_table,
            "schema_name": schema,
            "columns": columns,
            "row_count": row_count,
        })

        log.info("schema_manager.xlsx_sheet_ingested", table=full_table, rows=row_count)

    return results


# ---------------------------------------------------------------------------
# Schema introspection for LLM prompting
# ---------------------------------------------------------------------------

def get_project_schema_description(project_id: UUID) -> Optional[str]:
    """Get a human-readable schema description for a project's structured data.

    Returns None if no structured data exists.
    """
    try:
        conn = _get_duckdb()
        schema = _project_schema(project_id)

        # Check if schema exists
        schemas = conn.execute(
            "SELECT schema_name FROM information_schema.schemata WHERE schema_name = ?",
            [schema],
        ).fetchall()
        if not schemas:
            return None

        # Get all tables
        tables = conn.execute(
            "SELECT table_name FROM information_schema.tables WHERE table_schema = ?",
            [schema],
        ).fetchall()

        if not tables:
            return None

        lines = [f"Schema: {schema}", ""]
        for (table_name,) in tables:
            full_table = f"{schema}.{table_name}"
            lines.append(f"TABLE: {table_name}")

            # Get columns
            cols = conn.execute(
                """SELECT column_name, data_type
                   FROM information_schema.columns
                   WHERE table_schema = ? AND table_name = ?
                   ORDER BY ordinal_position""",
                [schema, table_name],
            ).fetchall()

            for col_name, col_type in cols:
                lines.append(f"  - {col_name}: {col_type}")

            # Get row count
            row_count = conn.execute(f"SELECT COUNT(*) FROM {full_table}").fetchone()[0]
            lines.append(f"  ({row_count} rows)")

            # Sample 3 rows for context
            sample = conn.execute(f"SELECT * FROM {full_table} LIMIT 3").fetchdf()
            if not sample.empty:
                lines.append(f"  Sample data: {sample.to_string(index=False, max_cols=8)}")
            lines.append("")

        return "\n".join(lines)

    except Exception as e:
        log.warning("schema_manager.introspection_failed", error=str(e))
        return None


def execute_sql(project_id: UUID, sql: str) -> dict:
    """Execute a read-only SQL query against a project's structured data.

    Returns:
        {"columns": [...], "rows": [...], "row_count": int, "truncated": bool}
    """
    conn = _get_duckdb()
    schema = _project_schema(project_id)

    # Set search path to project schema
    conn.execute(f"SET search_path TO '{schema}'")

    try:
        result = conn.execute(sql)
        columns = [desc[0] for desc in result.description]
        all_rows = result.fetchall()

        truncated = len(all_rows) > settings.sql_max_rows
        display_rows = all_rows[:settings.sql_max_rows]

        rows = [dict(zip(columns, row)) for row in display_rows]

        return {
            "columns": columns,
            "rows": rows,
            "row_count": len(all_rows),
            "truncated": truncated,
        }
    except Exception as e:
        return {
            "columns": [],
            "rows": [],
            "row_count": 0,
            "truncated": False,
            "error": str(e),
        }


def has_structured_data(project_id: UUID) -> bool:
    """Check if a project has any structured data tables."""
    try:
        conn = _get_duckdb()
        schema = _project_schema(project_id)
        tables = conn.execute(
            "SELECT COUNT(*) FROM information_schema.tables WHERE table_schema = ?",
            [schema],
        ).fetchone()
        return tables is not None and tables[0] > 0
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Type inference helpers
# ---------------------------------------------------------------------------

def _infer_column_types(headers: list[str], sample_rows: list[list[str]]) -> list[str]:
    """Infer DuckDB column types from sample data."""
    if not sample_rows:
        return ["VARCHAR"] * len(headers)

    types = []
    for col_idx in range(len(headers)):
        values = [
            row[col_idx] if col_idx < len(row) else ""
            for row in sample_rows
            if col_idx < len(row) and row[col_idx].strip()
        ]
        types.append(_infer_single_type(values))

    return types


def _infer_single_type(values: list[str]) -> str:
    """Infer the best DuckDB type for a column from sample values."""
    if not values:
        return "VARCHAR"

    # Try integer
    int_count = 0
    for v in values:
        try:
            int(v.replace(",", "").replace(" ", ""))
            int_count += 1
        except (ValueError, AttributeError):
            pass
    if int_count > len(values) * 0.8:
        return "BIGINT"

    # Try float / monetary
    float_count = 0
    for v in values:
        try:
            cleaned = v.replace(",", "").replace("$", "").replace("€", "").replace("£", "").replace(" ", "").replace("%", "")
            if cleaned:
                float(cleaned)
                float_count += 1
        except (ValueError, AttributeError):
            pass
    if float_count > len(values) * 0.8:
        return "DOUBLE"

    # Try date
    date_count = 0
    import re
    date_pattern = re.compile(r"\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4}|\d{2}-\d{2}-\d{4}")
    for v in values:
        if date_pattern.match(v.strip()):
            date_count += 1
    if date_count > len(values) * 0.8:
        return "DATE"

    return "VARCHAR"


def _cast_value(val: str, dtype: str):
    """Cast a string value to the appropriate Python type for DuckDB insertion."""
    if not val or not val.strip():
        return None

    val = val.strip()

    if dtype == "BIGINT":
        try:
            return int(val.replace(",", "").replace(" ", ""))
        except (ValueError, TypeError):
            return None

    if dtype == "DOUBLE":
        try:
            cleaned = val.replace(",", "").replace("$", "").replace("€", "").replace("£", "").replace("%", "").replace(" ", "")
            return float(cleaned)
        except (ValueError, TypeError):
            return None

    if dtype == "DATE":
        return val  # Let DuckDB handle date parsing

    return val
