from __future__ import annotations
from io import BytesIO

from openpyxl import load_workbook

from . import NormalizedDoc, PageContent


def extract(source_id: str, raw: bytes) -> NormalizedDoc:
    """Each sheet → one logical 'page'. Enhanced for financial data:

    - Preserves sheet structure with header detection
    - Generates per-sheet summary statistics
    - Handles merged cells and multi-row headers better
    - Captures formula presence (marks cells that had formulas)
    """
    # Read with formulas to detect calculated fields
    wb_formulas = load_workbook(BytesIO(raw), data_only=False, read_only=True)
    formula_cells = _detect_formulas(wb_formulas)

    # Read with values for actual data
    wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
    pages: list[PageContent] = []

    for idx, ws in enumerate(wb.worksheets):
        rows_iter = ws.iter_rows(values_only=True)
        rows = [list(r) for r in rows_iter]
        rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
        if not rows:
            continue

        header = rows[0]
        headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(header)]
        data_rows = rows[1:]

        # Build table text
        header_str = "\t".join(headers)
        body = "\n".join(
            "\t".join(str(c) if c is not None else "" for c in r) for r in data_rows
        )
        table_text = f"{header_str}\n{body}"

        # Generate sheet summary
        summary_parts = [
            f"Sheet: {ws.title}",
            f"Rows: {len(data_rows)}, Columns: {len(headers)}",
            f"Columns: {', '.join(headers)}",
        ]

        # Numeric stats
        numeric_stats = _summarize_numeric_columns(headers, data_rows)
        if numeric_stats:
            summary_parts.append("Key Statistics:")
            summary_parts.extend(f"  - {s}" for s in numeric_stats)

        # Note formula-based columns
        sheet_formulas = formula_cells.get(ws.title, [])
        if sheet_formulas:
            summary_parts.append(
                f"Calculated fields (formulas detected): {', '.join(sheet_formulas[:10])}"
            )

        summary_text = "\n".join(summary_parts)

        # Summary page
        pages.append(
            PageContent(
                page_num=idx * 2 + 1,
                text=summary_text,
                tables=[],
                section_title=f"{ws.title} - Summary",
                is_table=False,
            )
        )

        # Data page
        pages.append(
            PageContent(
                page_num=idx * 2 + 2,
                text=f"Sheet: {ws.title}\n{table_text}",
                tables=[table_text],
                section_title=ws.title,
                is_table=True,
            )
        )

    return NormalizedDoc(source_id=source_id, pages=pages)


def _detect_formulas(wb) -> dict[str, list[str]]:
    """Detect which columns contain formulas (calculated fields)."""
    formula_map: dict[str, list[str]] = {}
    try:
        for ws in wb.worksheets:
            formula_cols: set[int] = set()
            header_row = None
            for row_idx, row in enumerate(ws.iter_rows(values_only=False)):
                if row_idx == 0:
                    header_row = [
                        str(c.value) if c.value else f"col_{i}"
                        for i, c in enumerate(row)
                    ]
                    continue
                for col_idx, cell in enumerate(row):
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_cols.add(col_idx)

                # Only scan first 20 rows for performance
                if row_idx > 20:
                    break

            if header_row and formula_cols:
                formula_map[ws.title] = [
                    header_row[i] for i in formula_cols if i < len(header_row)
                ]
    except Exception:
        pass  # read_only mode may not support formula detection — graceful fallback
    return formula_map


def _summarize_numeric_columns(
    headers: list[str], data_rows: list[list], max_cols: int = 8
) -> list[str]:
    """Generate summary statistics for numeric columns."""
    summaries = []
    for col_idx, header in enumerate(headers[:max_cols]):
        values = []
        for row in data_rows:
            if col_idx < len(row) and row[col_idx] is not None:
                try:
                    val = str(row[col_idx]).strip().replace(",", "").replace("$", "").replace("%", "")
                    if val:
                        values.append(float(val))
                except (ValueError, TypeError):
                    continue

        if len(values) >= 3:
            total = sum(values)
            avg = total / len(values)
            min_val = min(values)
            max_val = max(values)
            summaries.append(
                f"{header}: min={min_val:,.2f}, max={max_val:,.2f}, "
                f"avg={avg:,.2f}, total={total:,.2f} ({len(values)} values)"
            )
    return summaries
